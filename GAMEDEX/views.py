# Módulo principal de vistas para GAMEDEX (usuarios, carrito, productos, facturación)
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User, Group
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib import messages
from django.contrib.auth import logout
from django.views.decorators.cache import never_cache
from django.http import HttpResponse, HttpResponseForbidden
from reportlab.pdfgen import canvas
from django.http import JsonResponse
from requests import post
from GAMEDEX.forms import EditarPerfilForm
from .models import Comentario, Comunidad, Perfil, Producto, Publicacion
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from .forms import ComunidadForm
from django.views.decorators.csrf import csrf_exempt



# =====================================
# INICIO
# =====================================


def api_productos(request):
    productos = Producto.objects.all()[:20]

    data = [
        {
            "id": p.id,
            "nombre": p.nombre,
            "precio": str(p.precio)
        }
        for p in productos
    ]

    return JsonResponse(data, safe=False)

def inicio(request):
    print("ENTRANDO A INICIO 🔥")

    productos = Producto.objects.filter(publicado=True, destacado=True)
    print("PRODUCTOS:", productos)

    return render(request, 'inicio.html', {'productos': productos})

# =====================================
# REGISTRO
# =====================================
def registro(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)   
        if form.is_valid():
            user = form.save()

            grupo, _ = Group.objects.get_or_create(name='Usuario')
            user.groups.add(grupo)

            messages.success(request, "Usuario registrado correctamente.")
            return redirect('login')
    else:
        form = UserCreationForm()

    return render(request, 'registro.html', {'form': form})


# =====================================
# DASHBOARD USUARIO
# =====================================
@login_required
@never_cache
def dashboard_usuario(request):

    productos = Producto.objects.filter(publicado=True)

    carrito = request.session.get("carrito", {})
    total_carrito = sum(item["cantidad"] for item in carrito.values())


    return render(request, "dashboard_usuario.html", {
        "productos": productos,
        "total_carrito": total_carrito
    })


# =====================================
# CARRITO
# =====================================
@login_required
def agregar_carrito(request, producto_id):

    producto = get_object_or_404(Producto, id=producto_id)

    carrito = request.session.get("carrito", {})

    if str(producto_id) in carrito:
        carrito[str(producto_id)]["cantidad"] += 1
    else:
        carrito[str(producto_id)] = {
            "nombre": producto.nombre,
            "precio": float(producto.precio),
            "cantidad": 1
        }

    request.session["carrito"] = carrito
    request.session.modified = True

    return redirect("dashboard_usuario")

def ver_carrito(request):

    carrito = request.session.get("carrito", {})
    productos_carrito = []
    total = 0

    for producto_id, item in carrito.items():

        producto = get_object_or_404(Producto, id=producto_id)

        cantidad = item["cantidad"]
        precio = float(item["precio"])

        subtotal = precio * cantidad
        total += subtotal

        productos_carrito.append({
            "producto": producto,
            "cantidad": cantidad,
            "subtotal": subtotal
        })

    return render(request, "carrito.html", {
        "productos_carrito": productos_carrito,
        "total": total
    })


def quitar_unidad(request, producto_id):

    carrito = request.session.get("carrito", {})

    if str(producto_id) in carrito:
        carrito[str(producto_id)]["cantidad"] -= 1

        if carrito[str(producto_id)]["cantidad"] <= 0:
            del carrito[str(producto_id)]

    request.session["carrito"] = carrito
    request.session.modified = True

    return redirect("ver_carrito")

def eliminar_producto(request, producto_id):

    carrito = request.session.get("carrito", {})

    if str(producto_id) in carrito:
        del carrito[str(producto_id)]

    request.session["carrito"] = carrito
    request.session.modified = True

    return redirect("ver_carrito")


# =====================================
# COMPRA
# =====================================

def comprar_carrito(request):

    carrito = request.session.get("carrito", {})

    if not carrito:
        messages.error(request, "El carrito está vacío.")
        return redirect("dashboard_usuario")

    productos_factura = []
    total = 0

    for producto_id, item in carrito.items():

        producto = get_object_or_404(Producto, id=producto_id)

        cantidad = item.get("cantidad", 0)
        precio = float(item.get("precio", 0))

        if cantidad > producto.cantidad:
            messages.error(request, f"No hay suficiente stock de {producto.nombre}")
            return redirect("ver_carrito")

        subtotal = precio * cantidad

        producto.cantidad -= cantidad
        producto.save()

        productos_factura.append({
            "nombre": producto.nombre,
            "descripcion": producto.descripcion,
            "precio": precio,
            "cantidad": cantidad,
            "subtotal": subtotal,

        })

        total += subtotal

    request.session["factura"] = {
        "productos": productos_factura,
        "total": total
    }

    request.session["carrito"] = {}

    request.session.modified = True
    request.session.save()

    return redirect("/factura/")

    # =====================================
    # 🔥 GUARDAR FACTURA (FORMA SEGURA)
    # =====================================
    request.session["factura"] = {
        "productos": productos_factura,
        "total": total
    }

    # limpiar carrito
    request.session["carrito"] = {}

    # 🔥 FORZAR GUARDADO REAL EN SESIÓN
    request.session.modified = True
    request.session.save()

    return redirect("factura")  

# =====================================
# FACTURA
# =====================================
def factura(request):

    factura = request.session.get("factura")

    if not factura:
        messages.error(request, "No hay factura disponible.")
        return redirect("dashboard_usuario")

    return render(request, "factura.html", {
        "productos": factura["productos"],
        "total": factura["total"],
        "usuario": request.user
    })


def descargar_factura_pdf(request):

    factura = request.session.get("factura")

    if not factura:
        return HttpResponse("No hay factura disponible")

    template = get_template("descargar_factura_pdf.html")

    html = template.render({
        "productos": factura["productos"],
        "total": factura["total"],
        "usuario": request.user
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="factura_gamedex.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("Error al generar PDF")

    return response

# =====================================
# PERFIL
# =====================================
def editar_perfil(request):

    user = request.user
    perfil = user.perfil

    if request.method == 'POST':
        user.username = request.POST.get('username')
        user.email = request.POST.get('email')

        perfil.telefono = request.POST.get('telefono')
        perfil.direccion = request.POST.get('direccion')

        user.save()
        perfil.save()

        messages.success(request, "Perfil actualizado correctamente")
        return redirect('dashboard_usuario')

    return render(request, 'editar_perfil.html', {
        'user': user,
        'perfil': perfil
    })


# =====================================
# DASHBOARD ADMIN
# =====================================
@login_required
@never_cache
def dashboard_admin(request):

    if not request.user.groups.filter(name="Administrador").exists():
        messages.error(request, "No tienes permisos.")
        return redirect("redireccion_dashboard")

    query = request.GET.get("q")
    usuarios_list = User.objects.all()

    if query:
        usuarios_list = usuarios_list.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query)
        )

    paginator = Paginator(usuarios_list.order_by("username"), 5)
    usuarios = paginator.get_page(request.GET.get("page"))

    # 🔥 CONTADORES
    total_usuarios = User.objects.count()
    total_vendedores = User.objects.filter(groups__name="Vendedor").count()
    total_admins = User.objects.filter(groups__name="Administrador").count()

    return render(request, "dashboard_admin.html", {
        "usuarios": usuarios,
        "query": query,
        "total_usuarios": total_usuarios,
        "total_vendedores": total_vendedores,
        "total_admins": total_admins
    })

def lista_usuarios(request):
    query = request.GET.get("q", "")
    rol = request.GET.get("rol", "")

    usuarios = User.objects.all()

    if query:
        usuarios = usuarios.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query)
        )

    if rol:
        usuarios = usuarios.filter(groups__name=rol)

    usuarios = usuarios.distinct()

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return render(request, "partials/tabla_usuarios.html", {
            "usuarios": usuarios
        })

    return render(request, "admin/usuarios.html", {
        "usuarios": usuarios,
        "total_usuarios": User.objects.count(),
        "total_vendedores": User.objects.filter(groups__name="Vendedor").count(),
        "total_admins": User.objects.filter(groups__name="Administrador").count(),
    })


# =====================================
# COMUNIDAD Y PUBLICACIONES
# =====================================

def admin_comunidades(request):
    if request.user.perfil.rol != "Administrador":
        return redirect('inicio')

    comunidades = Comunidad.objects.all()

    return render(request, 'admin_comunidades.html', {
        'comunidades': comunidades
    })

def ver_comunidad(request, id):
    comunidad = get_object_or_404(Comunidad, id=id)
    publicaciones = Publicacion.objects.filter(comunidad=comunidad).order_by('-id')

    if request.method == "POST":
        contenido = request.POST.get("contenido")
        imagen = request.FILES.get("imagen")
        if contenido or imagen:
            Publicacion.objects.create(
                contenido=contenido,
                imagen=imagen,
                autor=request.user,
                comunidad=comunidad
            )
        return redirect('ver_comunidad', id=comunidad.id)

    return render(request, "ver_comunidad.html", {
        "comunidad": comunidad,
        "publicaciones": publicaciones
    })

def crear_publicacion(request, id) -> JsonResponse:
    if request.method == "POST" and request.user.is_authenticated:
        comunidad = get_object_or_404(Comunidad, id=id)
        contenido = request.POST.get("contenido", "").strip()
        imagen = request.FILES.get("imagen", None)

        if contenido or imagen:
            publicacion = Publicacion.objects.create(
                comunidad=comunidad,
                autor=request.user,
                contenido=contenido,
                imagen=imagen
            )

            # Retornamos datos de la publicación como JSON
            data = {
                "id": publicacion.id,
                "autor": publicacion.autor.username,
                "contenido": publicacion.contenido,
                "imagen_url": publicacion.imagen.url if publicacion.imagen else "",
                "fecha": publicacion.created_at.strftime("%Y-%m-%d %H:%M"),
            }
            return JsonResponse(data)
        
    return JsonResponse({"error": "No se pudo crear la publicación"}, status=400)

def eliminar_publicacion(request, post_id):
    if request.method == "POST":
        post = get_object_or_404(Publicacion, id=post_id)
        user = request.user

        # Verificar permisos
        if user == post.autor or user == post.comunidad.creador or getattr(user.perfil, 'rol', '') == "Administrador":
            post.delete()
            return JsonResponse({"success": True})
        else:
            return JsonResponse({"success": False, "error": "No tienes permisos para eliminar esta publicación"})

    return JsonResponse({"success": False, "error": "Método no permitido"})


@login_required
def crear_publicacion_ajax(request, id):
    if request.method == "POST":
        comunidad = get_object_or_404(Comunidad, id=id)
        contenido = request.POST.get("contenido")
        imagen = request.FILES.get("imagen")
        if not contenido and not imagen:
            return JsonResponse({"error": "Debes escribir algo o subir una imagen."})
        
        publicacion = Publicacion.objects.create(
            contenido=contenido,
            imagen=imagen,
            autor=request.user,
            comunidad=comunidad
        )

        return JsonResponse({
            "id": publicacion.id,
            "contenido": publicacion.contenido,
            "autor": publicacion.autor.username,
            "imagen_url": publicacion.imagen.url if publicacion.imagen else ""
        })
    return JsonResponse({"error": "Método no permitido."})

def dar_like(request, id):
    post = get_object_or_404(Publicacion, id=id)

    if request.user in post.likes.all():
        post.likes.remove(request.user)
        liked = False
    else:
        post.likes.add(request.user)
        liked = True

    return JsonResponse({
        "liked": liked,
        "total_likes": post.total_likes()
    })

def comentar(request, id):
    post = get_object_or_404(Publicacion, id=id)

    if request.method == "POST":
        contenido = request.POST.get("contenido")

        Comentario.objects.create(
            publicacion=post,
            autor=request.user,
            contenido=contenido
        )

    return redirect('ver_comunidad', id=post.comunidad.id)

def lista_comunidades(request):
    comunidades = Comunidad.objects.all()

    es_admin = False
    if request.user.is_authenticated:
        es_admin = request.user.perfil.rol == "Administrador"

    return render(request, "comunidades.html", {
        "comunidades": comunidades,
        "es_admin": es_admin
    })

def crear_comunidad(request):
    if request.method == "POST":

        print("FILES:", request.FILES)  # 👈 AQUÍ
        print("POST:", request.POST)    # (opcional, para ver todo)

        nombre = request.POST.get("nombre")
        descripcion = request.POST.get("descripcion")
        imagen = request.FILES.get("imagen")

        print("Imagen recibida:", imagen)  # 👈 EXTRA DEBUG

        Comunidad.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            imagen=imagen,
            creador=request.user
        )

        return redirect('comunidades')

    return render(request, 'crear_comunidad.html')

def editar_comunidad(request, id):
    comunidad = get_object_or_404(Comunidad, id=id)

    # Verifica permisos: creador o admin
    if request.user != comunidad.creador and request.user.perfil.rol != "Administrador":
        return HttpResponseForbidden("No tienes permiso para editar esta comunidad.")

    if request.method == "POST":
        form = ComunidadForm(request.POST, request.FILES, instance=comunidad)
        if form.is_valid():
            form.save()
            return redirect('ver_comunidad', id=comunidad.id)
    else:
        form = ComunidadForm(instance=comunidad)

    return render(request, "editar_comunidad.html", {"form": form, "comunidad": comunidad})

def eliminar_comunidad(request, id):
    comunidad = get_object_or_404(Comunidad, id=id)

    if request.user != comunidad.creador and request.user.perfil.rol != "Administrador":
        return HttpResponseForbidden()

    comunidad.delete()
    return redirect('comunidades')

# =====================================
# INVENTARIO ADMIN
# =====================================
@login_required



def inventario_admin(request):
    productos = Producto.objects.all()

    query = request.GET.get("q")
    filtro = request.GET.get("filtro")

    # 🔎 BUSCADOR
    if query:
        productos = productos.filter(
            Q(nombre__icontains=query) |
            Q(descripcion__icontains=query)
        )

    # ⚠️ FILTROS
    if filtro == "stock_bajo":
        productos = productos.filter(cantidad__lt=5, cantidad__gt=0)

    elif filtro == "sin_stock":
        productos = productos.filter(cantidad=0)

    # 📄 PAGINACIÓN
    paginator = Paginator(productos, 10)
    page = request.GET.get("page")
    productos = paginator.get_page(page)

    return render(request, "admin/inventario.html", {
        "productos": productos,
        "query": query,
        "filtro": filtro
    })

#======================================
#exportar pdf de productos del admin
#=======================================
@login_required
def exportar_pdf_inventario(request):
    productos = Producto.objects.all()

    template = get_template("pdf_inventario.html")
    html = template.render({"productos": productos})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventario.pdf"'

    pisa.CreatePDF(html, dest=response)
    return response

#======================================
#exportar excel de productos del admin
#=======================================
@login_required
def exportar_excel_inventario(request):
    productos = Producto.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    ws.append(["ID", "Producto", "Descripción", "Precio", "Stock", "Vendedor"])

    for p in productos:
        ws.append([
            p.id,
            p.nombre,
            p.descripcion,
            float(p.precio),
            p.cantidad,
            p.vendedor.username
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=inventario.xlsx'

    wb.save(response)
    return response


# =====================================
# ELIMINAR PRODUCTO (ADMIN)
# =====================================
@login_required
def eliminar_producto_admin(request, id):

    if not request.user.groups.filter(name="Administrador").exists():
        messages.error(request, "No tienes permisos.")
        return redirect("redireccion_dashboard")

    producto = get_object_or_404(Producto, id=id)
    producto.delete()

    messages.success(request, "Producto eliminado correctamente.")

    return redirect("inventario_admin")
# =====================================
# CREAR USUARIO (ADMIN)
# =====================================
@login_required
def crear_usuario(request):

    if not request.user.groups.filter(name="Administrador").exists():
        messages.error(request, "No tienes permisos.")
        return redirect("redireccion_dashboard")

    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        rol = request.POST.get("rol")

        if User.objects.filter(username=username).exists():
            messages.error(request, "El usuario ya existe.")
            return redirect("crear_usuario")

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password
        )

        grupo, _ = Group.objects.get_or_create(name=rol)
        user.groups.add(grupo)

        messages.success(request, "Usuario creado correctamente.")
        return redirect("dashboard_admin")

    grupos = Group.objects.all()
    return render(request, "crear_usuario.html", {"grupos": grupos})


# =====================================
# EDITAR USUARIO (ADMIN)
# =====================================
@login_required
def editar_usuario(request, user_id):

    if not request.user.groups.filter(name="Administrador").exists():
        messages.error(request, "No tienes permisos.")
        return redirect("redireccion_dashboard")

    usuario = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        usuario.username = request.POST.get("username")
        usuario.email = request.POST.get("email")

        rol = request.POST.get("rol")
        nueva_password = request.POST.get("password")

        if nueva_password:
            usuario.set_password(nueva_password)

        usuario.save()

        usuario.groups.clear()
        grupo, _ = Group.objects.get_or_create(name=rol)
        usuario.groups.add(grupo)

        messages.success(request, "Usuario actualizado correctamente.")
        return redirect("dashboard_admin")

    grupos = Group.objects.all()

    return render(request, "editar_usuario.html", {
        "usuario": usuario,
        "grupos": grupos
    })
# =====================================
# exportar pdf de usuarios (ADMIN)
# =====================================
def exportar_pdf_usuarios(request):
    usuarios = User.objects.all()

    template = get_template("pdf_usuarios.html")
    html = template.render({"usuarios": usuarios})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="usuarios.pdf"'

    pisa.CreatePDF(html, dest=response)

    return response

# =====================================
# exportar excel de usuarios (ADMIN)
# =====================================
def exportar_excel_usuarios(request):
    usuarios = User.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    # Encabezados
    ws.append(["Usuario", "Email", "Rol"])

    for usuario in usuarios:
        roles = ", ".join([g.name for g in usuario.groups.all()])
        ws.append([
            usuario.username,
            usuario.email,
            roles if roles else "Sin rol"
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'

    wb.save(response)

    return response


# =====================================
# REGISTRO PUBLICO
# =====================================
def registro_publico(request):

    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        rol = request.POST.get("rol")  # El nombre que viene del select

        if User.objects.filter(username=username).exists():
            messages.error(request, "El usuario ya existe.")
            return redirect("registro_publico")

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password
        )

        # 🔥 ASIGNAR GRUPO SEGÚN ELECCIÓN
        if rol == "Vendedor":
            grupo, _ = Group.objects.get_or_create(name="Vendedor")
        else:
            grupo, _ = Group.objects.get_or_create(name="Cliente")

        user.groups.add(grupo)

        messages.success(request, "Cuenta creada correctamente.")
        return redirect("login")

    return render(request, "registro_publico.html")


# =====================================
# DASHBOARD VENDEDOR
# =====================================
@login_required
@never_cache


def dashboard_vendedor(request):

    if not request.user.groups.filter(name="Vendedor").exists():
        messages.error(request, "No tienes permiso.")
        return redirect("redireccion_dashboard")

    productos = Producto.objects.filter(vendedor=request.user).order_by('-id')

    # 🔎 BUSQUEDA
    q = request.GET.get('q')
    if q:
        productos = productos.filter(nombre__icontains=q)

    # 🔥 CONTADORES (ANTES de paginar)
    total_productos = productos.count()
    publicados = productos.filter(publicado=True).count()
    borradores = productos.filter(publicado=False).count()
    activos = productos.filter(cantidad__gt=1).count()

    # 📄 PAGINADOR (5 por página)
    paginator = Paginator(productos, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ⚡ AJAX (clave para que funcione tu dashboard pro)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, "partials/tabla_productos.html", {
            "page_obj": page_obj
        })

    return render(request, "dashboard_vendedor.html", {
        "page_obj": page_obj,   # 👈 IMPORTANTE
        "total_productos": total_productos,
        "publicados": publicados,
        "borradores": borradores,
        "activos": activos
    })

from django.template.loader import get_template
from xhtml2pdf import pisa
from django.http import HttpResponse

# =====================================
# Editar perfil (Vendedor)
# =====================================
 
@login_required
def editar_perfil(request):
    user = request.user  # 🔥 SOLO el usuario logueado

    if request.method == 'POST':
        form = EditarPerfilForm(request.POST, instance=user)

        if form.is_valid():
            user = form.save(commit=False)

            password = form.cleaned_data.get("password")

            if password:
                user.set_password(password)  # 🔥 encripta correctamente

            user.save()

            messages.success(request, "Perfil actualizado correctamente")

            return redirect("login")  # ⚠️ se cierra sesión al cambiar contraseña

    else:
        form = EditarPerfilForm(instance=user)

    return render(request, "editar_perfil.html", {"form": form})


# =====================================
# pdf de productos publicados del vendedor
# =====================================
@login_required
def exportar_pdf_vendedor(request):

    productos = Producto.objects.filter(
        vendedor=request.user,
        publicado=True
    )

    template = get_template("pdf_productos.html")

    html = template.render({
        "productos": productos,
        "usuario": request.user
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="productos.pdf"'

    pisa.CreatePDF(html, dest=response)

    return response
from openpyxl import Workbook

# =====================================
# excel de productos publicados del vendedor
# =====================================

@login_required
def exportar_excel_vendedor(request):

    productos = Producto.objects.filter(
        vendedor=request.user,
        publicado=True
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # encabezados
    ws.append(["Nombre", "Precio", "cantidad"])

    # datos
    for p in productos:
        ws.append([p.nombre, float(p.precio), p.cantidad])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'

    wb.save(response)

    return response



# =====================================
# CREAR PRODUCTO
# =====================================
@login_required
def crear_producto(request):

    if request.method == "POST":
        nombre = request.POST.get("nombre")
        descripcion = request.POST.get("descripcion")
        precio = request.POST.get("precio")
        cantidad = request.POST.get("cantidad")
        imagen = request.FILES.get("imagen")

        Producto.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            precio=precio,
            cantidad=cantidad,
            imagen=imagen,
            vendedor=request.user
        )

        messages.success(request, "Producto creado correctamente.")
        return redirect("dashboard_vendedor")

    return render(request, "crear_producto.html")

from django.shortcuts import get_object_or_404, redirect

# =====================================
# destacar producto (Vendedor)
# =====================================

def toggle_destacado(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    producto.destacado = not producto.destacado
    producto.save()

    return redirect('dashboard_vendedor')  # o donde estés mostrando productos

# =====================================
# EDITAR PRODUCTO
# =====================================
@login_required
def editar_producto(request, producto_id):

    producto = get_object_or_404(Producto, id=producto_id)

    if request.method == "POST":
        producto.nombre = request.POST.get("nombre")
        producto.descripcion = request.POST.get("descripcion")
        producto.precio = request.POST.get("precio")
        producto.cantidad = request.POST.get("cantidad")

        if request.FILES.get("imagen"):
            producto.imagen = request.FILES.get("imagen")

        producto.save()

        messages.success(request, "Producto actualizado.")
        return redirect("dashboard_vendedor")

    return render(request, "editar_producto.html", {"producto": producto})


# =====================================
# ELIMINAR PRODUCTO
# =====================================
@login_required
def eliminar_producto(request, producto_id):

    producto = get_object_or_404(Producto, id=producto_id)

    producto.delete()

    messages.success(request, "Producto eliminado.")
    return redirect("dashboard_vendedor")
# =====================================
# PUBLICAR / DESPUBLICAR PRODUCTO
# =====================================
@login_required
def toggle_publicacion(request, producto_id):

    producto = get_object_or_404(Producto, id=producto_id)

    # 🔐 Validación: solo el dueño puede cambiarlo
    if producto.vendedor != request.user:
        messages.error(request, "No tienes permiso para modificar este producto.")
        return redirect("dashboard_vendedor")

    # 🔄 Alternar estado
    producto.publicado = not producto.publicado
    producto.save()

    estado = "publicado" if producto.publicado else "ocultado"
    messages.success(request, f"Producto {estado} correctamente.")

    return redirect("dashboard_vendedor")


# =====================================
# REDIRECCIÓN
# =====================================
@login_required
def redireccion_dashboard(request):

    if request.user.groups.filter(name="Administrador").exists():
        return redirect("dashboard_admin")
    elif request.user.groups.filter(name="Vendedor").exists():
        return redirect("dashboard_vendedor")
    else:
        return redirect("dashboard_usuario")


# =====================================
# LOGOUT
# =====================================
@login_required
def cerrar_sesion(request):
    logout(request)
    return redirect("login")
